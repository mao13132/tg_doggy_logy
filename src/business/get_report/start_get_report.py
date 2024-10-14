# ---------------------------------------------
# Program by @developer_telegrams
#
#
# Version   Date        Info
# 1.0       2023    Initial Version
#
# ---------------------------------------------
import asyncio
import os
import time

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.cell import WriteOnlyCell

from settings import report_path
from src.business.get_report._get_question_by_user import get_question_by_user
from src.business.get_report.get_yers import get_years
from src.business.questions.questions import QUESTIONS
from src.utils.logger._logger import logger_msg


class StartGetReport:
    def __init__(self, settings):
        self.settings = settings

        self.bot = settings['bot']

        self.loop = settings['loop']

        self.id_user = settings['id_user']

        self.user_list = settings['user_list']

        self.question_list = settings['question_list']

        self.save_file_name = f"{report_path}{os.sep}{self.id_user}.xlsx"

        self.colums = ['Юзернейм', 'Имя', 'Фамилия', 'Премиум', 'Дата первого касания', 'Последняя активность']

    def write_data(self, ws, count_def, user, questions_user):

        print(f'Добавил в exel файл {count_def} строчек')

        row_cell = []

        cell1 = WriteOnlyCell(ws, value=user[2])
        cell2 = WriteOnlyCell(ws, value=user[3])
        cell3 = WriteOnlyCell(ws, value=user[4])
        cell4 = WriteOnlyCell(ws, value=user[5])
        cell5 = WriteOnlyCell(ws, value=user[6])
        cell6 = WriteOnlyCell(ws, value=user[7])

        row_cell.append(cell1)
        row_cell.append(cell2)
        row_cell.append(cell3)
        row_cell.append(cell4)
        row_cell.append(cell5)
        row_cell.append(cell6)

        for count_quest, quest in QUESTIONS.items():
            try:
                user_answer = questions_user[count_quest]['answer']
            except:
                user_answer = f'Не заполнил'

            row_cell.append(WriteOnlyCell(ws, value=user_answer))

        years = get_years(questions_user)

        row_cell.append(WriteOnlyCell(ws, value=years))

        ws.append(row_cell)

        return True

    def create_title(self, ws):

        global_count = 0

        row_cell = []

        other_columns = [value['title'] for key, value in QUESTIONS.items()]

        all_columns = self.colums + other_columns + ['Лет']

        for count, col in enumerate(all_columns):
            cell = WriteOnlyCell(ws, value=col)

            cell.font = Font(bold=True)

            row_cell.append(cell)

            global_count += 1

        ws.append(row_cell)

        return True

    async def iter_rows(self, ws):
        count_def = 3

        for count_member, user in enumerate(self.user_list):
            id_user = user[1]

            questions_user = get_question_by_user(id_user, self.question_list)

            try:
                write_data = self.write_data(ws, count_def, user, questions_user)
            except Exception as es:
                logger_msg(f'StartGetReport: Исключение {es}')

            count_def += 1

        return True

    async def one_sheet(self, ws):
        res_create_title = self.create_title(ws)

        res_iter = await self.iter_rows(ws)

        return True

    async def _start_get_report(self):

        wb = Workbook(write_only=True)

        ws = wb.create_sheet()

        result = await self.one_sheet(ws)

        try:

            wb.save(self.save_file_name)

            wb.close()

        except Exception as es:
            error_ = f'Ошибка при сохранение отчета excel "{es}"'

            logger_msg(error_)

            return False

        print(f'Сохранил отчет')

        return self.save_file_name

    async def send_report_file(self):

        try:

            with open(self.save_file_name, 'rb') as file:
                await self.bot.send_document(self.id_user, document=file,
                                             caption=f'Отчет на {len(self.user_list)} человек(а)')

        except Exception as es:
            error_ = f'Не могу выслать файл отчета. Причина: "{es}"'

            logger_msg(error_)

            return False

        return True

    def start_get_report(self):

        try:
            res_report = asyncio.run(self._start_get_report())
        except Exception as es:
            error_ = f'Ошибка при формирование отчета: "{es}"'

            logger_msg(error_)

            return False

        try:
            res_report = asyncio.run_coroutine_threadsafe(self.send_report_file(), self.loop)
        except Exception as es:
            error_ = f'Ошибка при формирование отчета: "{es}"'

            logger_msg(error_)

            return False

        return res_report
